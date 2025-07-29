# Library
import json
import logging
from pathlib import Path
from unittest import result
from openpyxl import load_workbook

from utils.constant import FOLDER_INPUT, FOLDER_OUTPUT


def import_json(path: str) -> any:
    """
    Import data from a JSON file.

    Args:
        file_path: The path to the JSON file.

    Return the parsed JSON data.
    """
    file_path = Path(path)

    result: any = None
    if not file_path.exists():
        logging.error(f'ERROR_JSON import_json("{file_path}") : {err}')
        return result

    with open(file_path, "r", encoding="utf-8") as f:
        try:
            result = json.load(f)
        except json.JSONDecodeError as e:
            logging.error(
                f'ERROR_JSON_INVALID import_json("{file_path}") : {e.doc, e.pos}'
            )

    print(f'JSON_IMPORTED "{result}"')
    return result


def export_json(
    data: any,
    filename: str | None = None,
    path: str | None = None,
    indent: int | None = None,
) -> str:

    if path:
        file_path = path
    else:
        # Create output directory if it doesn't exist
        Path(FOLDER_OUTPUT).mkdir(parents=True, exist_ok=True)
        file_path = f"{FOLDER_OUTPUT}/{filename}"

    # Convert data to JSON string if it isn't already
    if not isinstance(data, str):
        data = json.dumps(data, indent=indent, default=str)

    with open(file_path, "w", encoding="utf-8") as f:
        f.write(data)

    print(f'JSON_EXPORTED "{file_path}" data:\n{data}')

    return file_path


def extract_urls_from_excel(filename: str):
    """
    Extracts all URLs from an Excel (.xlsx) file

    Args:
        file_path (str): Path to the Excel file

    Returns:
        list: Unique list of URLs found in the workbook
    """

    # Create output directory if it doesn't exist

    input_file = f"{FOLDER_INPUT}/{filename}"
    print(f'Read XLSX "{input_file}"')

    wb = load_workbook(filename=input_file)
    urls = set()

    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]

        for row in sheet.iter_rows():
            for cell in row:
                # Check for explicit hyperlinks
                if cell.hyperlink:
                    urls.add(cell.hyperlink.target)

                # Check for text that looks like a URL
                if isinstance(cell.value, str) and cell.value.startswith(
                    ("http://", "https://")
                ):
                    urls.add(cell.value.strip())

    result = list(urls)
    print(f'Found "{len(urls)}" URLs :')
    # for index, url in enumerate(result):
    #     print(f"{index}. {url}")

    return result
